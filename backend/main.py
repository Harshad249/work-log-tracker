"""
WorkLog Tracker — FastAPI Backend
Delta Metrology Internal Tool
"""
import os
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))
from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
from fastapi import FastAPI, HTTPException, Depends, status
app = FastAPI()
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from typing import Optional, List
from datetime import datetime, date, timedelta
from jose import JWTError, jwt
from passlib.context import CryptContext
import databases, sqlalchemy
import io, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─── CONFIG ─────────────────────────────────────────────────────────────────

SECRET_KEY   = os.getenv("SECRET_KEY", "change-me-in-production-please")
ALGORITHM    = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 8   # 8-hour sessions

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./worklog.db")

# ─── DATABASE ────────────────────────────────────────────────────────────────

database  = databases.Database(DATABASE_URL)
metadata  = sqlalchemy.MetaData()

users_table = sqlalchemy.Table("users", metadata,
    sqlalchemy.Column("id",         sqlalchemy.String,  primary_key=True),
    sqlalchemy.Column("name",       sqlalchemy.String,  nullable=False),
    sqlalchemy.Column("email",      sqlalchemy.String,  unique=True, nullable=False),
    sqlalchemy.Column("hashed_pw",  sqlalchemy.String,  nullable=False),
    sqlalchemy.Column("role",       sqlalchemy.String,  default="employee"),
    sqlalchemy.Column("created_at", sqlalchemy.DateTime, default=datetime.utcnow),
)
from sqlalchemy import create_engine
engine = create_engine(DATABASE_URL)
metadata.create_all(engine)

logs_table = sqlalchemy.Table("work_logs", metadata,
    sqlalchemy.Column("id",         sqlalchemy.String,  primary_key=True),
    sqlalchemy.Column("user_id",    sqlalchemy.String,  nullable=False),
    sqlalchemy.Column("user_name",  sqlalchemy.String,  nullable=False),
    sqlalchemy.Column("date",       sqlalchemy.Date,    nullable=False),
    sqlalchemy.Column("project",    sqlalchemy.String,  nullable=False),
    sqlalchemy.Column("task",       sqlalchemy.Text,    nullable=False),
    sqlalchemy.Column("hours",      sqlalchemy.Float,   nullable=False),
    sqlalchemy.Column("remarks",    sqlalchemy.String,  default=""),
    sqlalchemy.Column("created_at", sqlalchemy.DateTime, default=datetime.utcnow),
    sqlalchemy.Column("updated_at", sqlalchemy.DateTime, default=datetime.utcnow),
)

engine = sqlalchemy.create_engine(DATABASE_URL.replace("postgresql://", "postgresql+psycopg2://"))
metadata.create_all(engine)

# ─── AUTH HELPERS ─────────────────────────────────────────────────────────────

pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")

def hash_password(pw: str) -> str:
    return pwd_ctx.hash(pw)

def verify_password(pw: str, hashed: str) -> bool:
    return pwd_ctx.verify(pw, hashed)

def create_access_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(token: str = Depends(oauth2_scheme)):
    creds_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if not user_id:
            raise creds_exception
    except JWTError:
        raise creds_exception

    row = await database.fetch_one(users_table.select().where(users_table.c.id == user_id))
    if not row:
        raise creds_exception
    return dict(row)

def require_admin(user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

# ─── PYDANTIC SCHEMAS ─────────────────────────────────────────────────────────

class UserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str
    role: str = "employee"

class UserOut(BaseModel):
    id: str
    name: str
    email: str
    role: str
    created_at: datetime

class LogCreate(BaseModel):
    date: date
    project: str
    task: str
    hours: float
    remarks: Optional[str] = ""

class LogUpdate(BaseModel):
    project: Optional[str]
    task: Optional[str]
    hours: Optional[float]
    remarks: Optional[str]

class LogOut(BaseModel):
    id: str
    user_id: str
    user_name: str
    date: date
    project: str
    task: str
    hours: float
    remarks: str
    created_at: datetime

class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserOut

# ─── APP ─────────────────────────────────────────────────────────────────────

app = FastAPI(title="WorkLog API", version="1.0.0", docs_url="/docs")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],          # Tighten in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup():
    await database.connect()

@app.on_event("shutdown")
async def shutdown():
    await database.disconnect()

# ─── AUTH ROUTES ─────────────────────────────────────────────────────────────

import uuid

@app.post("/auth/register", response_model=Token, status_code=201)
async def register(body: UserCreate):
    existing = await database.fetch_one(users_table.select().where(users_table.c.email == body.email))
    if existing:
        raise HTTPException(400, "Email already registered")

    uid = str(uuid.uuid4())
    await database.execute(users_table.insert().values(
        id=uid, name=body.name, email=body.email,
        hashed_pw=hash_password(body.password), role=body.role,
        created_at=datetime.utcnow()
    ))
    user = await database.fetch_one(users_table.select().where(users_table.c.id == uid))
    token = create_access_token({"sub": uid})
    return {"access_token": token, "token_type": "bearer", "user": dict(user)}


@app.post("/auth/login", response_model=Token)
async def login(form: OAuth2PasswordRequestForm = Depends()):
    user = await database.fetch_one(users_table.select().where(users_table.c.email == form.username))
    if not user or not verify_password(form.password, user["hashed_pw"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_access_token({"sub": user["id"]})
    return {"access_token": token, "token_type": "bearer", "user": dict(user)}


@app.get("/auth/me", response_model=UserOut)
async def me(user=Depends(get_current_user)):
    return user

# ─── WORK LOG ROUTES ─────────────────────────────────────────────────────────

@app.post("/logs", response_model=LogOut, status_code=201)
async def create_log(body: LogCreate, user=Depends(get_current_user)):
    if body.hours <= 0 or body.hours > 24:
        raise HTTPException(400, "Hours must be between 0.5 and 24")

    # Employees can only log for today
    if user["role"] == "employee" and body.date != date.today():
        raise HTTPException(400, "Employees can only log work for today")

    # Check duplicate for same user + date
    existing = await database.fetch_one(
        logs_table.select().where(
            (logs_table.c.user_id == user["id"]) & (logs_table.c.date == body.date)
        )
    )
    if existing:
        raise HTTPException(400, "Entry already exists for this date. Use PUT to update.")

    lid = str(uuid.uuid4())
    now = datetime.utcnow()
    await database.execute(logs_table.insert().values(
        id=lid, user_id=user["id"], user_name=user["name"],
        date=body.date, project=body.project, task=body.task,
        hours=body.hours, remarks=body.remarks or "",
        created_at=now, updated_at=now
    ))
    return await database.fetch_one(logs_table.select().where(logs_table.c.id == lid))


@app.get("/logs", response_model=List[LogOut])
async def get_logs(
    user_id: Optional[str] = None,
    project: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user=Depends(get_current_user)
):
    q = logs_table.select()

    # Employees only see their own logs
    if current_user["role"] == "employee":
        q = q.where(logs_table.c.user_id == current_user["id"])
    elif user_id:
        q = q.where(logs_table.c.user_id == user_id)

    if project:
        q = q.where(logs_table.c.project == project)
    if date_from:
        q = q.where(logs_table.c.date >= date_from)
    if date_to:
        q = q.where(logs_table.c.date <= date_to)

    q = q.order_by(logs_table.c.date.desc())
    rows = await database.fetch_all(q)
    return [dict(r) for r in rows]


@app.put("/logs/{log_id}", response_model=LogOut)
async def update_log(log_id: str, body: LogUpdate, user=Depends(get_current_user)):
    log = await database.fetch_one(logs_table.select().where(logs_table.c.id == log_id))
    if not log:
        raise HTTPException(404, "Log not found")

    # Employees can only edit their own logs, and only on the same day
    if user["role"] == "employee":
        if log["user_id"] != user["id"]:
            raise HTTPException(403, "Not your log")
        if log["date"] != date.today():
            raise HTTPException(400, "Employees can only edit today's entries")

    update_data = {k: v for k, v in body.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()

    await database.execute(logs_table.update().where(logs_table.c.id == log_id).values(**update_data))
    return await database.fetch_one(logs_table.select().where(logs_table.c.id == log_id))


@app.delete("/logs/{log_id}", status_code=204)
async def delete_log(log_id: str, user=Depends(get_current_user)):
    log = await database.fetch_one(logs_table.select().where(logs_table.c.id == log_id))
    if not log:
        raise HTTPException(404, "Log not found")
    if user["role"] != "admin" and log["user_id"] != user["id"]:
        raise HTTPException(403, "Not authorized")
    await database.execute(logs_table.delete().where(logs_table.c.id == log_id))

# ─── REPORTS & EXPORTS ───────────────────────────────────────────────────────

@app.get("/reports/monthly/xlsx")
async def export_monthly_xlsx(year: int, month: int, _=Depends(require_admin)):
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    date_from = date(year, month, 1)
    date_to   = date(year, month, last_day)

    rows = await database.fetch_all(
        logs_table.select()
        .where((logs_table.c.date >= date_from) & (logs_table.c.date <= date_to))
        .order_by(logs_table.c.date, logs_table.c.user_name)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{str(month).zfill(2)}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a3a6e")
    headers = ["Employee Name", "Date", "Project", "Task Description", "Hours Worked", "Remarks"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row["user_name"])
        ws.cell(row=row_idx, column=2, value=row["date"].strftime("%d %b %Y"))
        ws.cell(row=row_idx, column=3, value=row["project"])
        ws.cell(row=row_idx, column=4, value=row["task"])
        ws.cell(row=row_idx, column=5, value=row["hours"])
        ws.cell(row=row_idx, column=6, value=row["remarks"] or "")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 44
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 22

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"WorkLog_{year}_{str(month).zfill(2)}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/reports/monthly/csv")
async def export_monthly_csv(year: int, month: int, _=Depends(require_admin)):
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    date_from = date(year, month, 1)
    date_to   = date(year, month, last_day)

    rows = await database.fetch_all(
        logs_table.select()
        .where((logs_table.c.date >= date_from) & (logs_table.c.date <= date_to))
        .order_by(logs_table.c.date, logs_table.c.user_name)
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Employee Name", "Date", "Project", "Task Description", "Hours Worked", "Remarks"])
    for row in rows:
        writer.writerow([
            row["user_name"], row["date"].strftime("%d %b %Y"),
            row["project"], row["task"], row["hours"], row["remarks"] or ""
        ])

    output.seek(0)
    filename = f"WorkLog_{year}_{str(month).zfill(2)}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/reports/summary")
async def monthly_summary(year: int, month: int, _=Depends(require_admin)):
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    date_from = date(year, month, 1)
    date_to   = date(year, month, last_day)

    rows = await database.fetch_all(
        logs_table.select()
        .where((logs_table.c.date >= date_from) & (logs_table.c.date <= date_to))
    )

    by_employee, by_project = {}, {}
    for r in rows:
        by_employee.setdefault(r["user_name"], 0)
        by_employee[r["user_name"]] += r["hours"]
        by_project.setdefault(r["project"], 0)
        by_project[r["project"]] += r["hours"]

    return {
        "year": year, "month": month,
        "total_entries": len(rows),
        "total_hours": round(sum(r["hours"] for r in rows), 2),
        "by_employee": [{"name": k, "hours": round(v, 2)} for k, v in sorted(by_employee.items(), key=lambda x: -x[1])],
        "by_project":  [{"project": k, "hours": round(v, 2)} for k, v in sorted(by_project.items(), key=lambda x: -x[1])],
    }

# ─── USER MANAGEMENT (ADMIN) ─────────────────────────────────────────────────

@app.get("/users", response_model=List[UserOut])
async def list_users(_=Depends(require_admin)):
    return [dict(r) for r in await database.fetch_all(users_table.select().order_by(users_table.c.name))]


@app.post("/users", response_model=UserOut, status_code=201)
async def create_user(body: UserCreate, _=Depends(require_admin)):
    existing = await database.fetch_one(users_table.select().where(users_table.c.email == body.email))
    if existing:
        raise HTTPException(400, "Email already registered")
    uid = str(uuid.uuid4())
    await database.execute(users_table.insert().values(
        id=uid, name=body.name, email=body.email,
        hashed_pw=hash_password(body.password), role=body.role,
        created_at=datetime.utcnow()
    ))
    return await database.fetch_one(users_table.select().where(users_table.c.id == uid))


@app.delete("/users/{user_id}", status_code=204)
async def delete_user(user_id: str, current_user=Depends(require_admin)):
    if user_id == current_user["id"]:
        raise HTTPException(400, "Cannot delete yourself")
    await database.execute(users_table.delete().where(users_table.c.id == user_id))


@app.get("/health")
async def health():
    return {"status": "ok", "service": "worklog-api"}
print("DATABASE_URL:", DATABASE_URL)